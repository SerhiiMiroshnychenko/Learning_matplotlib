import control
import numpy as np
import matplotlib.pyplot as plt
import tbcontrol
import scipy.optimize
from tbcontrol.responses import fopdt, sopdt
from ipywidgets import interact
from openpyxl import Workbook

tbcontrol.expectversion("0.1.10")
Greal = control.TransferFunction([1, 2], [2, 3, 4, 1])
ts, ys = control.step_response(Greal)
yinitial = 10
measurement_noise = np.random.randn(len(ys)) * 0.05
ym = ys + yinitial + measurement_noise
print(f'{ts=}')
print(f'{ym=}')

def show_1():
    plt.scatter(ts, ym)
    plt.plot(ts, ys + yinitial, color='red')
    plt.show()


def resultplot(K, tau, theta, y0):
    plt.scatter(ts, ym)
    plt.plot(ts, fopdt(ts, K, tau, theta, y0), color='red')
    plt.show()

interact(resultplot,
         K=(1., 10.),
         tau=(0., 10.),
         theta=(0., 10.),
         y0=(0., 20.))

[K, tau, theta, y0], _ = scipy.optimize.curve_fit(fopdt, ts, ym, [2, 4, 1, 10])
print('Коефіцієнти K, tau, theta, y0', [K, tau, theta, y0])

[K_2, tau_2, zeta_2, theta_2, y0_2], _ = scipy.optimize.curve_fit(sopdt, ts, ym, [2, 2, 1.5, 1, 10])
print('Коефіцієнти K_2, tau_2, zeta_2, theta_2, y0_2', [K_2, tau_2, zeta_2, theta_2, y0_2])


def show_2():
    plt.figure(figsize=(10, 5))
    plt.scatter(ts, ym, label='Data')
    global fopdt_data
    fopdt_data = fopdt(ts, K, tau, theta, y0)
    print(f'{fopdt_data=}')
    print(type(fopdt_data))
    global sopdt_data
    sopdt_data = sopdt(ts, K_2, tau_2, zeta_2, theta_2, y0_2)
    print(f'{sopdt_data=}')
    plt.plot(ts, fopdt_data, color='red', label='FOPDT fit')
    plt.plot(ts, sopdt_data, color='green', label='SOPDT fit')

    plt.plot(ts, ys + 10, color='blue', label='Original')
    plt.legend(loc='best')
    plt.show()

def write_in_excel():
    # Створення нового робочого зошита
    workbook = Workbook()
    sheet = workbook.active

    # Запис назв списків у перший рядок
    sheet.cell(row=1, column=1, value='ts')
    sheet.cell(row=1, column=2, value='ym')
    sheet.cell(row=1, column=3, value='fopdt')
    sheet.cell(row=1, column=4, value='sopdt')

    # Запис даних у стовпці
    for i, value in enumerate(ts, start=2):
        sheet.cell(row=i, column=1, value=value)

    for i, value in enumerate(ym, start=2):
        sheet.cell(row=i, column=2, value=value)

    for i, value in enumerate(fopdt_data, start=2):
        sheet.cell(row=i, column=3, value=value)

    for i, value in enumerate(sopdt_data, start=2):
        sheet.cell(row=i, column=4, value=value)

    # Збереження зошита
    workbook.save('items-original.xlsx')


if __name__ == "__main__":
    show_1()
    show_2()
    write_in_excel()