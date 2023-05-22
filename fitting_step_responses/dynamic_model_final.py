import scipy.optimize
from tbcontrol.responses import fopdt, sopdt
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook


# Зчитування даних з файлу Excel
data = pd.read_excel('data.xlsx', header=None)
ts = data.iloc[0].tolist()  # Перший рядок - ts
ym = data.iloc[1].tolist()  # Другий рядок - ym

[K, tau, theta, y0], _ = scipy.optimize.curve_fit(fopdt, ts, ym, [2, 4, 1, 10])
print('Коефіцієнти K, tau, theta, y0', [K, tau, theta, y0])

[K_2, tau_2, zeta_2, theta_2, y0_2], _ = scipy.optimize.curve_fit(sopdt, ts, ym, [2, 2, 1.5, 1, 10])
print('Коефіцієнти K_2, tau_2, zeta_2, theta_2, y0_2', [K_2, tau_2, zeta_2, theta_2, y0_2])


plt.style.use('Solarize_Light2')
fig, ax = plt.subplots()
ax.scatter(ts, ym, color='steelblue', label='Data')
fopdt_data = fopdt(ts, K, tau, theta, y0)
print(f'{fopdt_data=}')
ax.plot(ts, fopdt_data, color='orange', label='FOPDT fit')
sopdt_data = sopdt(ts, K_2, tau_2, zeta_2, theta_2, y0_2)
print(f'{sopdt_data=}')
ax.plot(ts, sopdt_data, color='gold', label='SOPDT fit')
try:
    plt.legend(loc='best')
except Exception as e:
    print(e.__class__, e)
ax.set(xlabel='Часові моменти ts', ylabel='Відповідь системи',
       title='Залежність № 2')
ax.grid(True)


def write_in_excel():
    # Створення нового робочого зошита
    workbook = Workbook()
    sheet = workbook.active

    # Запис назв списків у перший рядок
    sheet.cell(row=1, column=1, value='ts')
    sheet.cell(row=1, column=2, value='ym')
    sheet.cell(row=1, column=3, value='fopdt')
    sheet.cell(row=1, column=4, value='sopdt')

    # Запис даних у стовпці
    for i, value in enumerate(ts, start=2):
        sheet.cell(row=i, column=1, value=value)

    for i, value in enumerate(ym, start=2):
        sheet.cell(row=i, column=2, value=value)

    for i, value in enumerate(fopdt_data, start=2):
        sheet.cell(row=i, column=3, value=value)

    for i, value in enumerate(sopdt_data, start=2):
        sheet.cell(row=i, column=4, value=value)

    # Збереження зошита
    workbook.save('items-real.xlsx')


if __name__ == "__main__":
    fig.savefig("plot_2.png")
    plt.show()
    write_in_excel()
